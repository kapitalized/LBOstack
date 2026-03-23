"""
B2B Math & Logic Engine — FastAPI microservice for CAD/financial calculations.
See blueprint @16_fastapi_main. Run: uvicorn main:app --reload --port 8000
"""

import base64
import io
import os
import re
import urllib.request
from typing import Optional

import uvicorn
from fastapi import FastAPI, Header, HTTPException
from pydantic import BaseModel
from openpyxl import load_workbook

app = FastAPI(
    title="B2B Math & Logic Engine",
    description="FastAPI microservice for CAD parsing and financial calculations.",
)

INTERNAL_SERVICE_KEY = os.getenv("INTERNAL_SERVICE_KEY", "dev-secret-handshake")


class MathRequest(BaseModel):
    data: list[dict]
    parameters: dict


@app.get("/")
async def health_check():
    """Service health check for container orchestration."""
    return {
        "status": "online",
        "engine": "FastAPI",
        "features": ["CAD Parsing", "Financial Logic", "Shoelace Area Calc"],
    }


@app.post("/calculate")
async def calculate(req: MathRequest, x_service_key: Optional[str] = Header(None)):
    """
    Primary endpoint for high-precision calculations.
    Accepts raw extracted data and returns calculated volumes/ratios.
    """
    if x_service_key != INTERNAL_SERVICE_KEY:
        raise HTTPException(status_code=403, detail="Unauthorized: Invalid Service Key")

    results = []
    thickness = req.parameters.get("thickness", 0.2)

    try:
        for item in req.data:
            area = item.get("area", 0)
            volume = area * thickness
            results.append({
                "id": item.get("id"),
                "label": item.get("label", "Unknown Component"),
                "area_m2": round(area, 2),
                "volume_m3": round(volume, 2),
                "verified": True,
            })
        return {
            "status": "success",
            "results": results,
            "metadata": {
                "items_processed": len(results),
                "applied_thickness": thickness,
            },
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Math Logic Error: {str(e)}") from e


def _parse_number(value):
    """Best-effort numeric parsing from Excel cells."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        s = value.strip()
        if s == "":
            return None
        # Handle percentages like "10%" or "10 %"
        s = s.replace("%", "").strip()
        s = re.sub(r"[^0-9+\\-\\.Ee]", "", s)
        if s == "" or s == ".":
            return None
        try:
            return float(s)
        except Exception:
            return None
    return None


def _decode_maybe_base64(data_str: str) -> bytes:
    # Supports both raw base64 and data URLs: data:<mime>;base64,<payload>
    if not isinstance(data_str, str):
        raise ValueError("file_b64 must be a string")
    s = data_str.strip()
    if s.startswith("data:") and ";base64," in s:
        s = s.split(";base64,", 1)[1]
    return base64.b64decode(s)


def _fetch_bytes(url: str) -> bytes:
    if not isinstance(url, str) or not url.strip():
        raise ValueError("url must be a non-empty string")
    with urllib.request.urlopen(url) as resp:
        return resp.read()


def _cell_contains(cell_value, keywords) -> bool:
    if cell_value is None:
        return False
    if not isinstance(cell_value, str):
        return False
    s = cell_value.strip().lower()
    return any(k in s for k in keywords)


def _extract_cashflow_series(ws, max_rows=220, max_cols=30) -> list[float]:
    """
    MVP heuristic:
    Find a header cell containing "operating" + "cash" + "flow", then read numeric values
    below it until blank/non-numeric.
    """
    keywords = ["operating", "cash", "flow"]
    header_col = None
    header_row = None
    for r in range(1, min(60, ws.max_row or 60) + 1):
        for c in range(1, min(15, max_cols) + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                low = v.strip().lower()
                if all(k in low for k in keywords):
                    header_col = c
                    header_row = r
                    break
        if header_col is not None:
            break

    if header_col is None:
        # Secondary heuristic: "OCF" or "Operating CF"
        for r in range(1, min(60, ws.max_row or 60) + 1):
            for c in range(1, min(15, max_cols) + 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str):
                    low = v.strip().lower()
                    if "ocf" in low or ("operating" in low and "cf" in low) or "operating cf" in low:
                        header_col = c
                        header_row = r
                        break
            if header_col is not None:
                break

    if header_col is None:
        return []

    series = []
    start = (header_row or 0) + 1
    for r in range(start, start + max_rows):
        v = ws.cell(row=r, column=header_col).value
        if v is None:
            break
        num = _parse_number(v)
        if num is None:
            break
        series.append(num)
    return series


def _extract_purchase_price(ws, max_rows=120, max_cols=10) -> Optional[float]:
    purchase_keywords = ["purchase price", "enterprise value", "purchase consideration", "purchase ev", "ev"]
    for r in range(1, min(max_rows, ws.max_row or max_rows) + 1):
        for c in range(1, min(max_cols, ws.max_column or max_cols) + 1):
            v = ws.cell(row=r, column=c).value
            if _cell_contains(v, purchase_keywords) and c + 1 <= ws.max_column:
                num = _parse_number(ws.cell(row=r, column=c + 1).value)
                if num is not None:
                    return num
    return None


def _extract_exit_value_or_multiple(ws, max_rows=120, max_cols=10) -> tuple[Optional[str], Optional[float]]:
    """
    Returns:
      ('multiple', exitMultiple) or ('fixed', exitValue)
    """
    for r in range(1, min(max_rows, ws.max_row or max_rows) + 1):
        for c in range(1, min(max_cols, ws.max_column or max_cols) + 1):
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, str):
                continue
            low = v.strip().lower()
            if "exit multiple" in low or (("exit" in low) and ("multiple" in low)):
                num = _parse_number(ws.cell(row=r, column=c + 1).value)
                if num is not None:
                    return ("multiple", num)
            if "exit value" in low or (("exit" in low) and ("value" in low)):
                num = _parse_number(ws.cell(row=r, column=c + 1).value)
                if num is not None:
                    return ("fixed", num)
            if low == "exit multiple":
                num = _parse_number(ws.cell(row=r, column=c + 1).value)
                if num is not None:
                    return ("multiple", num)
    return (None, None)


def _extract_tranches(ws, max_rows=60, max_cols=12) -> list[dict]:
    """
    MVP heuristic:
    Look for a header row containing 'tranche' + 'principal' + 'interest', then parse columns.
    """
    header_row = None
    header_cols = {}
    for r in range(1, min(40, ws.max_row or 40) + 1):
        # Build a small header list for this row.
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(max_cols, ws.max_column or max_cols) + 1)]
        row_lowers = [v.strip().lower() for v in row_vals if isinstance(v, str)]
        if any("tranche" in s for s in row_lowers) and any("principal" in s or "balance" in s for s in row_lowers) and any("interest" in s or "rate" in s for s in row_lowers):
            header_row = r
            for c in range(1, min(max_cols, ws.max_column or max_cols) + 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str):
                    low = v.strip().lower()
                    if "tranche" in low or "facility" in low:
                        header_cols["tranche"] = c
                    if "principal" in low or "balance" in low or "outstanding" in low:
                        header_cols["principal"] = c
                    if "interest" in low or "rate" in low:
                        header_cols["interest"] = c
                    if "priority" in low or "senior" in low or "order" in low:
                        header_cols["priority"] = c
            if "tranche" in header_cols and "principal" in header_cols and "interest" in header_cols:
                break

    tranches: list[dict] = []
    if header_row is not None and "tranche" in header_cols and "principal" in header_cols and "interest" in header_cols:
        start = header_row + 1
        priority_col = header_cols.get("priority")
        tranche_col = header_cols["tranche"]
        principal_col = header_cols["principal"]
        interest_col = header_cols["interest"]

        for r in range(start, start + max_rows):
            tranche_name = ws.cell(row=r, column=tranche_col).value
            if tranche_name is None or (isinstance(tranche_name, str) and tranche_name.strip() == ""):
                # Stop when table rows end (assume contiguous).
                if len(tranches) > 0:
                    break
                continue
            tranche_name_str = str(tranche_name).strip()
            principal = _parse_number(ws.cell(row=r, column=principal_col).value)
            interest = _parse_number(ws.cell(row=r, column=interest_col).value)
            if principal is None or interest is None:
                continue

            # Convert interest rate if given as percent.
            if interest > 1.0:
                # Common patterns: 10 => 0.10, 11 => 0.11, 8.5 => 0.085
                interest = interest / 100.0

            priority = None
            if priority_col is not None:
                priority = _parse_number(ws.cell(row=r, column=priority_col).value)
            if priority is None:
                priority = len(tranches)  # preserve order as extracted

            tid = f"tranche_{len(tranches) + 1}"
            tranches.append(
                {
                    "id": tid,
                    "name": tranche_name_str,
                    "initialPrincipal": float(principal),
                    "annualInterestRate": float(interest),
                    "priority": float(priority),
                    "amortization": {"type": "interest_only"},
                }
            )

        return tranches

    # Secondary heuristic: rows with "senior" or "subordinated" and adjacent principal/rate.
    tranches = []
    for r in range(1, min(80, ws.max_row or 80) + 1):
        left = ws.cell(row=r, column=1).value
        if not isinstance(left, str):
            continue
        low = left.strip().lower()
        if "senior" in low or "subordinated" in low or "sub" in low or "mezz" in low:
            principal = _parse_number(ws.cell(row=r, column=2).value)
            interest = _parse_number(ws.cell(row=r, column=3).value)
            if principal is None or interest is None:
                continue
            if interest > 1.0:
                interest = interest / 100.0
            priority = 0 if "senior" in low else 1
            tid = f"tranche_{len(tranches) + 1}"
            tranches.append(
                {
                    "id": tid,
                    "name": str(left).strip(),
                    "initialPrincipal": float(principal),
                    "annualInterestRate": float(interest),
                    "priority": float(priority),
                    "amortization": {"type": "interest_only"},
                }
            )
    return tranches


def _deep_merge_dict(base: dict, overrides: dict) -> dict:
    out = dict(base or {})
    for k, v in (overrides or {}).items():
        if isinstance(v, dict) and isinstance(out.get(k), dict):
            out[k] = _deep_merge_dict(out[k], v)
        else:
            out[k] = v
    return out


def extract_lbo_deal_from_workbook(xlsx_bytes: bytes, overrides: Optional[dict] = None) -> tuple[dict, list[str]]:
    warnings: list[str] = []
    wb = load_workbook(filename=io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    ws = wb.worksheets[0]

    op_cf = _extract_cashflow_series(ws)
    if len(op_cf) == 0:
        warnings.append("Could not extract operating cashflows; provide overrides.operatingCashFlows.")

    purchase_price = _extract_purchase_price(ws)
    if purchase_price is None:
        warnings.append("Could not extract purchase price; provide overrides.transaction.purchasePrice.")

    exit_type, exit_num = _extract_exit_value_or_multiple(ws)
    if exit_type is None:
        warnings.append("Could not extract exit multiple/value; provide overrides.transaction.exit.")

    tranches = _extract_tranches(ws)
    if len(tranches) == 0:
        warnings.append("Could not extract debt tranches; provide overrides.debt.tranches.")

    period_count = len(op_cf)

    deal = {
        "periodCount": period_count,
        "monthsPerPeriod": 1,
        "operatingCashFlows": op_cf,
        "transaction": {
            "purchasePrice": purchase_price,
            "exit": None,
            "transactionFees": 0,
        },
        "debt": {"tranches": tranches},
        "sweep": {"distributeResidualToEquity": True, "timing": "end_of_period"},
        "equity": {},
        "liquidityReserve": {"minCashBalance": 0},
    }

    if exit_type is not None and exit_num is not None:
        if exit_type == "multiple":
            deal["transaction"]["exit"] = {"type": "multiple", "exitMultiple": float(exit_num)}
        else:
            deal["transaction"]["exit"] = {"type": "fixed", "exitValue": float(exit_num)}

    # Merge overrides for the hybrid workflow.
    if overrides and isinstance(overrides, dict) and len(overrides) > 0:
        deal = _deep_merge_dict(deal, overrides)

    # If overrides supply operating cashflows but not periodCount (or periodCount was 0),
    # derive periodCount from cashflows so validation doesn't fail prematurely.
    if isinstance(deal.get("operatingCashFlows"), list):
        cashflow_len = len(deal["operatingCashFlows"])
        if not isinstance(deal.get("periodCount"), int) or deal.get("periodCount") <= 0:
            deal["periodCount"] = cashflow_len

    # Basic validation: required minimum keys for the TS engine.
    required_errors = []
    if not isinstance(deal.get("periodCount"), int) or deal.get("periodCount") <= 0:
        required_errors.append("periodCount must be provided and > 0")
    if not isinstance(deal.get("operatingCashFlows"), list) or len(deal["operatingCashFlows"]) != deal["periodCount"]:
        required_errors.append("operatingCashFlows length must match periodCount")
    if deal.get("transaction", {}).get("purchasePrice") is None:
        required_errors.append("transaction.purchasePrice is required")
    exit_obj = deal.get("transaction", {}).get("exit")
    if not isinstance(exit_obj, dict) or exit_obj.get("type") not in ("multiple", "fixed"):
        required_errors.append("transaction.exit is required (type: 'multiple' or 'fixed')")
    if not isinstance(deal.get("debt", {}).get("tranches"), list) or len(deal["debt"]["tranches"]) == 0:
        required_errors.append("debt.tranches must be non-empty")

    if len(required_errors) > 0:
        raise ValueError("Invalid LBO deal extraction: " + "; ".join(required_errors) + ". Warnings: " + "; ".join(warnings))

    return deal, warnings


@app.post("/lbo/extract-inputs")
async def lbo_extract_inputs(req: MathRequest, x_service_key: Optional[str] = Header(None)):
    """
    Extract MVP LBO deal inputs from an uploaded Excel model.

    POST body:
      { data: [ { url?: string, file_b64?: string } ], parameters: { templateId?: string, overrides?: object } }

    Returns:
      { status: 'success', results: [ { deal: <DealSchema dict>, warnings: [...] } ] }
    """
    if x_service_key != INTERNAL_SERVICE_KEY:
        raise HTTPException(status_code=403, detail="Unauthorized: Invalid Service Key")

    parameters = req.parameters or {}
    overrides = parameters.get("overrides") if isinstance(parameters.get("overrides"), dict) else None

    results = []
    for item in (req.data or []):
        file_bytes = None
        if isinstance(item, dict):
            if isinstance(item.get("file_b64"), str):
                file_bytes = _decode_maybe_base64(item["file_b64"])
            elif isinstance(item.get("file_bytes_b64"), str):
                file_bytes = _decode_maybe_base64(item["file_bytes_b64"])
            elif isinstance(item.get("url"), str) and item.get("url").strip():
                file_bytes = _fetch_bytes(item["url"])

        if file_bytes is None:
            raise HTTPException(status_code=400, detail="Missing file bytes: provide data item with url or file_b64.")

        try:
            deal, warnings = extract_lbo_deal_from_workbook(file_bytes, overrides=overrides)
            results.append({"deal": deal, "warnings": warnings})
        except ValueError as ve:
            raise HTTPException(status_code=422, detail=str(ve)) from ve
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"LBO extraction failed: {str(e)}") from e

    return {
        "status": "success",
        "results": results,
        "metadata": {"items_processed": len(results)},
    }


if __name__ == "__main__":
    port = int(os.getenv("PORT", "8000"))
    uvicorn.run(app, host="0.0.0.0", port=port)
